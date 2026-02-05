import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from django.http import HttpResponse
from datetime import datetime
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch

class ExportService:
    @staticmethod
    def export_to_excel(data, headers, filename="report", user_name=None):
        """
        Generates a professional Excel file with styling.
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Report"

        # Define styles
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        cell_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        border = Border(
            left=Side(style='thin', color='D3D3D3'),
            right=Side(style='thin', color='D3D3D3'),
            top=Side(style='thin', color='D3D3D3'),
            bottom=Side(style='thin', color='D3D3D3')
        )

        # Add report info at the top
        info_font = Font(size=10, italic=True, color="666666")
        current_row = 1
        
        if user_name:
            ws.cell(row=current_row, column=1, value=f"Exported by: {user_name}")
            ws.cell(row=current_row, column=1).font = info_font
            current_row += 1
        
        ws.cell(row=current_row, column=1, value=f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}")
        ws.cell(row=current_row, column=1).font = info_font
        current_row += 2  # Add space before table

        # Write and style header
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col_num, value=header.replace('_', ' ').title())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Alternate row colors for better readability
        light_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Write data with styling
        start_data_row = current_row + 1
        for row_num, row_data in enumerate(data, start_data_row):
            for col_num, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                
                # Convert complex types to readable strings
                if isinstance(value, (list, dict)):
                    value = str(value)
                elif isinstance(value, datetime):
                    value = value.strftime("%Y-%m-%d %H:%M")
                
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.alignment = cell_alignment
                cell.border = border
                
                # Alternate row colors
                if row_num % 2 == 0:
                    cell.fill = light_fill

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column].width = adjusted_width

        # Freeze the header row
        ws.freeze_panes = f"A{current_row + 1}"

        # Prepare response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response["Content-Disposition"] = f'attachment; filename="{filename}_{timestamp}.xlsx"'
        
        wb.save(response)
        return response

    @staticmethod
    def export_to_csv(data, headers, filename="report", user_name=None):
        """
        Generates a CSV file with UTF-8 BOM for Excel compatibility.
        """
        import csv
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        response['Content-Disposition'] = f'attachment; filename="{filename}_{timestamp}.csv"'

        # Add UTF-8 BOM for proper Excel opening
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # Add report info
        if user_name:
            writer.writerow([f"Exported by: {user_name}"])
        writer.writerow([f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}"])
        writer.writerow([])  # Empty row for spacing
        
        # Write header with nice formatting
        header_row = [h.replace('_', ' ').title() for h in headers]
        writer.writerow(header_row)
        
        for row in data:
            row_data = []
            for header in headers:
                val = row.get(header, "")
                if isinstance(val, (list, dict)):
                    val = str(val)
                elif isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d %H:%M")
                row_data.append(val)
            writer.writerow(row_data)
            
        return response

    @staticmethod
    def export_to_pdf(data, headers, filename="report", title=None, user_name=None):
        """
        Generates a professional PDF report with table styling.
        """
        try:
            response = HttpResponse(content_type='application/pdf')
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            response['Content-Disposition'] = f'attachment; filename="{filename}_{timestamp}.pdf"'
            
            # Create the PDF document
            doc = SimpleDocTemplate(response, pagesize=A4,
                                   rightMargin=30, leftMargin=30,
                                   topMargin=30, bottomMargin=30)
            
            elements = []
            styles = getSampleStyleSheet()
            
            # Add custom title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#4472C4'),
                spaceAfter=30,
                alignment=1  # Center
            )
            
            # Add title
            report_title = title or filename.replace('_', ' ').title()
            elements.append(Paragraph(report_title, title_style))
            
            # Add user and timestamp info
            info_style = ParagraphStyle(
                'InfoStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                alignment=1,
                spaceAfter=5
            )
            
            if user_name:
                elements.append(Paragraph(f"Exported by: {user_name}", info_style))
            
            elements.append(Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %H:%M')}", info_style))
            elements.append(Spacer(1, 20))
            
            # Prepare table data
            table_data = []
            
            # Add headers
            header_row = [h.replace('_', ' ').title() for h in headers]
            table_data.append(header_row)
            
            # Add data rows
            for row in data:
                row_data = []
                for header in headers:
                    value = row.get(header, "")
                    if isinstance(value, (list, dict)):
                        value = str(value)
                    elif isinstance(value, datetime):
                        value = value.strftime("%Y-%m-%d %H:%M")
                    row_data.append(str(value))
                table_data.append(row_data)
            
            # Create table
            table = Table(table_data)
            
            # Style the table
            table_style = TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows styling
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('TOPPADDING', (0, 1), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                
                # Alternate row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F2F2F2')]),
            ])
            
            table.setStyle(table_style)
            elements.append(table)
            
            # Add footer
            elements.append(Spacer(1, 30))
            footer_style = ParagraphStyle(
                'Footer',
                parent=styles['Normal'],
                fontSize=8,
                textColor=colors.grey,
                alignment=1
            )
            elements.append(Paragraph(f"Total Records: {len(data)}", footer_style))
            
            # Build PDF
            doc.build(elements)
            return response
            
        except ImportError:
            # Fallback if reportlab is not installed
            response = HttpResponse(
                "PDF generation requires 'reportlab' library. Please install it: pip install reportlab",
                content_type='text/plain'
            )
            return response